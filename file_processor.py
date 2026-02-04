"""
Advanced File Processing for Threat Modeling
Supports PDF, DOCX, XLSX, images (PNG, JPG) with OCR
"""

import io
import base64
from typing import Optional, Dict, Any
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

def extract_text_from_pdf(file_content: bytes) -> str:
    """Extract text from PDF file"""
    try:
        import PyPDF2
        pdf_file = io.BytesIO(file_content)
        reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"PDF extraction error: {e}")
        return f"[Error extracting PDF: {str(e)}]"


def extract_text_from_docx(file_content: bytes) -> str:
    """Extract text from DOCX file"""
    try:
        from docx import Document
        doc_file = io.BytesIO(file_content)
        doc = Document(doc_file)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text.strip()
    except Exception as e:
        logger.error(f"DOCX extraction error: {e}")
        return f"[Error extracting DOCX: {str(e)}]"


def extract_text_from_xlsx(file_content: bytes) -> str:
    """Extract text from XLSX file"""
    try:
        import openpyxl
        xlsx_file = io.BytesIO(file_content)
        workbook = openpyxl.load_workbook(xlsx_file)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n## Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"XLSX extraction error: {e}")
        return f"[Error extracting XLSX: {str(e)}]"


def extract_text_from_image_ocr(file_content: bytes) -> str:
    """Extract text from image using OCR (Tesseract)"""
    try:
        from PIL import Image
        import pytesseract
        image = Image.open(io.BytesIO(file_content))
        text = pytesseract.image_to_string(image)
        return text.strip()
    except Exception as e:
        logger.error(f"OCR extraction error: {e}")
        return f"[Error extracting text from image: {str(e)}. Note: Tesseract OCR may not be installed.]"


def encode_image_for_claude(file_content: bytes, file_type: str) -> Dict[str, Any]:
    """Encode image for Claude Vision API"""
    try:
        # Determine media type
        media_type_map = {
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'gif': 'image/gif',
            'webp': 'image/webp'
        }
        media_type = media_type_map.get(file_type.lower(), 'image/jpeg')
        
        # Encode to base64
        base64_image = base64.standard_b64encode(file_content).decode('utf-8')
        
        return {
            'type': 'image',
            'source': {
                'type': 'base64',
                'media_type': media_type,
                'data': base64_image
            }
        }
    except Exception as e:
        logger.error(f"Image encoding error: {e}")
        return None


def process_file(filename: str, file_content: bytes, use_vision_api: bool = False, max_chars_per_file: int = 100000) -> str:
    """
    Process any file type and extract text content
    
    Args:
        filename: Name of the file
        file_content: Binary content of the file
        use_vision_api: If True, return image data for Claude Vision API instead of OCR
        max_chars_per_file: Maximum characters to extract per file (default 100k = ~25k tokens)
        
    Returns:
        Extracted text content or placeholder
    """
    try:
        file_extension = Path(filename).suffix.lower().lstrip('.')
        
        # Text files - read directly
        if file_extension in ['txt', 'md', 'csv', 'json', 'xml', 'log']:
            content = file_content.decode('utf-8', errors='ignore')
            return truncate_content(content, max_chars_per_file)
        
        # PDF files
        elif file_extension == 'pdf':
            content = extract_text_from_pdf(file_content)
            return truncate_content(content, max_chars_per_file)
        
        # Word documents
        elif file_extension in ['docx', 'doc']:
            if file_extension == 'docx':
                content = extract_text_from_docx(file_content)
                return truncate_content(content, max_chars_per_file)
            else:
                return f"[.DOC format not supported. Please convert to .DOCX]"
        
        # Excel files
        elif file_extension in ['xlsx', 'xls']:
            if file_extension == 'xlsx':
                content = extract_text_from_xlsx(file_content)
                return truncate_content(content, max_chars_per_file)
            else:
                return f"[.XLS format not supported. Please convert to .XLSX]"
        
        # Image files
        elif file_extension in ['png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp']:
            if use_vision_api:
                # Return special marker for vision API processing
                return f"[IMAGE_FOR_VISION_API: {filename}]"
            else:
                # Use OCR
                ocr_text = extract_text_from_image_ocr(file_content)
                content = f"### Image: {filename}\n[OCR Extracted Text]\n{ocr_text}"
                return truncate_content(content, max_chars_per_file)
        
        # Unsupported formats
        else:
            return f"[{file_extension.upper()} Document: {filename}]"
            
    except Exception as e:
        logger.error(f"File processing error for {filename}: {e}")
        return f"[Error processing {filename}: {str(e)}]"


def truncate_content(content: str, max_chars: int) -> str:
    """
    Truncate content to stay within character limits
    """
    if len(content) > max_chars:
        truncated = content[:max_chars]
        return f"{truncated}\n\n... [Content truncated - Original: {len(content):,} chars, Showing: {max_chars:,} chars]"
    return content
